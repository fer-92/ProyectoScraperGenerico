"""
Con este diseño el scrip se debería lanzar de esta manera:
python3 scrap.py 1
python3 scrap.py 2
python3 scrap.py 3
python3 scrap.py 4
python3 scrap.py 5
python3 scrap.py 6
python3 scrap.py 7
python3 scrap.py 8
python3 scrap.py 9
"""
import hashlib
import ast
from jsonmerge import merge
import datetime
import json
import pickle
import urllib
import requests
import csv
import re
import os
from unidecode import unidecode
from bs4 import BeautifulSoup
from openpyxl import Workbook
from nube import texto2els
import sys
from store import store
id_grupo_telegram = ""
def corresponde_procesar_id(k, terminacion_id):
    if k[-1] == terminacion_id:
        return True
    else:
        return False

    """
    :param k:
        es el id del grupo de telegram que que devuelve el json que tiene la sonda llamadao json_sonda2.bin
    :secreto:
        Se fija que terminanción de id_telegram tiene asiganada esta asignada esta instancia correr.
        Este dato se lo pasa como parámettro cuando se ejecuta.
        Por ejemplo: si tiene el parámetro 1 quire decir que el k tiene que terminar en 1 para devolver true,
        caso contrario devuelve false. Ejemplo: si el k es -123456789, termina en 9, etonces devuelve false, porque
        el parámetro es 1.
    :return:
        True or False
    """
def carga_portales(provincia):
    provincia = provincia.replace(" ","_")
    f = open("./LinksDeLasProvincias/"+provincia+".json", "r")
    ff = json.loads(f.read())
    return ff
def formar_json_portales(lista_de_provincias):
    base = None
    result = []
    result2 = {}
    for l in lista_de_provincias:
        #with open("./LinksDeLasProvincias/"+l+".json", "r") as infile:
           # links = infile.read().replace("\n","")
            #result.append(links)
        base = merge(base, carga_portales(l))
        result.extend(base["link"])
    return result
j_link_enviado = {}
j_pag_ne = {}
mugre = ["xmlns=http://www.w3.org/1999/>", "<\n", "\n>", "<<p>", "<p>", "</p", "xmlns=http://www.w3.org/1999/>",
         "xmlns=http://www.w3.org/1999/>", "<br />", "CDATA", "</div>>", "<div>", "</div>", "%>", "<iframe>",
         "</iframe>", "100%", "<div", "http://w3.org/", "xmlms", "xhtml", ";>", "<", ">", "'", '"', "\/", "]", "[",
         "/","-","ttp",":","swww"]
def configuracionExcels(url):
    global PagNoticiaLink
    PagNoticiaLink = Workbook()
    global hoja
    hoja = PagNoticiaLink.active
    hoja['A1'] = "URL"
    hoja['A2'] = url
    hoja['A3'] = "TEXTO OBTENIDO"
    hoja['B3'] = "LINKS DE LAS NOTICIAS"
    hoja['C3'] = "TAG CON EL QUE ENCONTRO LAS NOTICIAS"
    hoja['D3'] = "TODOS LOS LINKS DE LA NOTICIA"
    hoja['E3'] = "HTML"
    hoja['F3'] = "TITULO NOTICIA"
    hoja['G3'] = "DESCRIPCION NOTICIA"
    hoja['H3'] = "FECHA PUBLICACION NOTICIA"
    hoja['I3'] = "FECHA MODIFICACION NOTICIA"
    hoja['J3'] = "ERRORES"
    hoja['K3'] = "ERRORES"
    hoja['L3'] = "ERRORES"
    hoja['M3'] = "ERRORES"
    hoja['N3'] = "ERRORES"
    hoja['O3'] = "ERRORES"
    hoja['P3'] = "ERRORES"
    hoja['Q3'] = "ERRORES"
    return PagNoticiaLink, hoja
def log(texto):
    l = open("log.csv", "a")
    l.write(texto + "\n")
    l.close()

def contarElementosLista(lista):
    """
    Recibe una lista, y devuelve un diccionario con todas las repeticiones decada valor
    """
    return {i: lista.count(i) for i in lista}
#def save_persist(elem):
    #try:
        #vpath = "./persist"+nrogrupo+"/"

        #varchivo = vpath + elem + ".bin"
        #with open(varchivo, "bw") as archivo:
            #pickle.dump(eval(elem), archivo)
    #except Exception as e:

        #print("Except de save_persist", e)
def save_persist_dinamico(elem,id_grupo_telegram):
    try:
        vpath = "./persist/"

        varchivo = vpath + elem + id_grupo_telegram+".bin"
        with open(varchivo, "bw") as archivo:
            pickle.dump(eval(elem), archivo)
    except Exception as e:

        print("Except de save_persist", e)
def load_persist(elem):
    try:
        vpath = "./persist/"
        varchivo =  elem +".bin"
        with open(varchivo, "br") as archivo:
            # #print(pickle.load(archivo))
            return pickle.load(archivo)
    except Exception as e:
        print("269 - Except load_persit ", e)

def replaceBase(texto):
    texto = texto.replace("http:", " ").replace("//", " ").replace(".", " ").replace("www", " ").replace(
            "https:", " ").replace("/", " ").replace("\n", " ").replace("-", " ")
    return  texto
def replaceURL(texto):
    texto = texto.replace("http:", "").replace("//", "").replace(".", "").replace("www", "").replace(
            "https:", "").replace("/", "").replace("\n", "").replace("-", "")
    return  texto
def filtroReplace(object):
    object.replace("/", "").replace(":", "").replace("%", "").replace("-", "").replace("[", "").replace("]","").replace("<","").replace(">", "").replace("!", "").replace(",", "")
    return " ".join(object.split())
def filtro_repetida(j_i):

    try:

        dd = j_i['link'].replace("\n", "")[1:200]
        dd = limpiar(dd, mugre)
        dd = hashear(dd)


        r = False

        if dd in db_noticias2.keys():
            if (db_noticias2[dd] == 1):
                r = True
            else:
                r = False



        if not r:
            db_noticias2[dd] = 1
            print("\n ********* \n No encontrado: \n",dd,"\n",j_i['link'],"\n*************")
            save_persist_dinamico('db_noticias2',id_grupo_telegram)
        return r


    except Exception as e:
        print("329 - ", e)
def filtro_tema(j_i, tema):
    c1 = tema.upper() in j_i['desc'].upper()
    for j in tema.split(","):
        c2 = j.upper() in j_i['desc'].upper()
        c1 = c1 or c2
    if c1:
        r = True
    else:
        r = False
    return r
def filtro_tema2(texto, tema):
    # c1 = tema.upper() in texto.upper()
    r = False
    for j in tema:
        if j.upper() in texto.upper():
            r = True
        #c1 = c1 or c2
    if r:
        r = True
    else:
        r = False
    return r

    # for j in tema.split(","):
    #    c2 = j.upper() in texto.upper()
    #    c1 = c1 or c2
    # if c1:
    #    r = True
    # else:
    #    r = False
    # return r
def link2medio(link):
    r = ""
    try:
        r  = re.search( '(?<=www.)\w+', link ).group()
    except:
        try:
            r = re.search( '(?<=//)\w+', link ).group()
        except:
            r =""
            return r
    return r
class RSSParser(object):

    def parse(self, confiTagPage, url, tema):
        ListaDeLinks = []
        items = []
        Noticia = []
        RedesSociales = ["facebook", "twitter", "whatsapp"]
        self.url = url
        global urlCortada
        if "news.google" in url:
            urlCortada = "Google"
        else:

            urlCortada = replaceURL(url)
        LINKS = open("./LINKS/" + urlCortada + ".csv", "a", encoding='utf-8')

        configuracionExcels(url)
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:55.0) Gecko/20100101 Firefox/55.0',
            }
            response = requests.get(url, headers=headers).text
        except Exception as e:
            print("Error 1 - Obtener Response ", e)
        fila = 0
        try:
            LINKS.write('----------LINK-----------' + '\n' + str(url) + '\n')
            for Noti in confiTagPage["j"]["BuscarNoticia"]:
                try:
                    Noticias = eval(Noti)
                    fila += 1
                    if Noticias != []:
                        LINKS.write('----------EVALUADOR-----------:' + '\n' + str(Noti) + '\n')
                        for i, Noticiae in enumerate(Noticias):
                            texto = filtroReplace(Noticiae.get_text())
                            print(texto)
                            if filtro_tema2(texto, tema) and texto != '':
                                Noticia.append(Noticiae)
                                LIIINKS = [a['href'] for a in Noticiae.find_all('a', href=True)]
                                # LIIINKS = ', '+'\n'.join(LIIINKS)
                                print(LIIINKS)
                                LINKS.write(
                                    '----------HTML-----------:' + '\n' + str(filtroReplace(Noticiae.text)) + '\n')
                                LINKS.write('----------LINK-----------:' + '\n' + str(LIIINKS) + '\n')
                                hoja.cell(row=i + 4, column=1).value = filtroReplace(filtroReplace(Noticiae.text))
                                hoja.cell(row=i + 4, column=2).value = str(LIIINKS)
                                hoja.cell(row=i + 4, column=3).value = str(Noti)
                except Exception as e:

                    print("Error 2 - Obtener Articulos de noticias ", e)

            PagNoticiaLink.save('./Excel/' + urlCortada + '-Noticias.xlsx')
        except Exception as e:
            print("Error 3 - Obtener Articulos de noticias ", e)
        temp9 = ""
        row = 3
        PagNoticiaLink.save('./Excel/' + urlCortada + '-Noticias.xlsx')
        try:
            for i in Noticia:
                row += 1

                hoja.cell(row=row, column=5).value = str(i)
                texto = filtroReplace(i.get_text())
                if filtro_tema2(texto, tema) and texto != '':

                    ListaDeLinks = eval(confiTagPage["j"]["path"])
                    if ListaDeLinks != "":
                        SetDeLinks = set(ListaDeLinks)
                        url2 = url
                        try:
                            resultado = contarElementosLista(ListaDeLinks)
                            if resultado != {}:
                                maximo = max(resultado, key=resultado.get)
                                print("El valor mas repetido es el ", maximo, " con ", resultado[maximo], " veces")

                        except Exception as e:
                            print(" 4 - Obtener Resultado maximo de links ", e)
                            hoja.cell(row=fila + 4, column=12).value = str(e)
                            hoja.cell(row=fila + 4, column=13).value = str(resultado)
                            hoja.cell(row=fila + 4, column=14).value = str(maximo)
                        if len(SetDeLinks) == 1 and list(SetDeLinks)[0] != url:
                            temp9 = list(SetDeLinks)
                            temp9 = str(temp9[0])
                            hoja.cell(row=row, column=2).value = temp9
                            LINKS.write('----------LINK-----------' + '\n' + str(
                                unidecode(temp9)) + '\n' + '----------HTML-----------:' + '\n' + unidecode(
                                str(texto)) + '\n')
                        else:
                            if resultado != {}:
                                if resultado[maximo] >= 2:
                                    temp9 = maximo
                                    hoja.cell(row=row, column=2).value = temp9
                                    LINKS.write('----------LINK-----------' + '\n' + str(unidecode(
                                        temp9)) + '\n' + '----------HTML-----------:' + '\n' + unidecode(str(
                                        texto)) + '\n')
                                else:
                                    palabras = texto.split()
                                    palabras.append(tema)
                                    try:
                                        for l in ListaDeLinks:
                                            linkCortado = replaceBase(l).split()
                                            totalPalabras = len(
                                                [palabra for palabra in palabras if palabra in linkCortado])
                                            if 2 <= totalPalabras <= 20:
                                                totalRedes = len([redSocial for redSocial in RedesSociales if
                                                                  redSocial in l.lower()])
                                                if not totalRedes >= 1:
                                                    temp9 = l
                                                    LINKS.write(unidecode(temp9) + '\n')

                                    except Exception as e:
                                        hoja.cell(row=fila + 4, column=15).value = str(e.args)
                                        hoja.cell(row=fila + 4, column=16).value = str(i)
                                        print(" ERROR 5 - NO ESCRAPEO NADA ", e)
                                        print(" ********* URL no parseada correctamente: \n", url, "\n")
                                        print(i)
                                        print("**********************************************************")

                                        if not i.text in j_pag_ne.keys():
                                            j_pag_ne[i.text] = 1
                                            log("***** \n No scrapeó esta página: \n" + url + '\n' + str(
                                                i) + '\n ********')

                            """
                            print("- url: ", url)
                            print("- temp9:" , temp9)
                            print("--------------------------------")
                        """
                    if temp9[:1] == "/":
                        temp9 = temp9[1:]
                    if temp9[:5] == "jujuy":
                        temp9 = temp9.replace("jujuy","")
                    if temp9[:2] == "./":
                        temp9 = temp9[2:]
                    if temp9[:8] == "noticias":
                        temp9 = temp9[8:]
                    if "http" in temp9:
                        url2 = temp9
                        temp9 = ""

                        # print(" url final:  ", url +  temp9 + temp10 + temp11, "\n temp9: ", temp9, "\n temp10: ",temp10, "\n temp11: ",temp11, "\n temp12: ",temp12)
                    if "news.google" in url2:
                        url2 = "https://news.google.com/"
                    j_i = {"link": url2 + temp9,
                           "desc": texto,
                           "tmpItems": i}
                    try:
                        if not filtro_repetida(j_i):
                            archivoCSV = []
                            LinkNotcia = j_i["link"]
                            DescripcionNoticia = ""
                            fechaPublicacion = ""
                            response2 = requests.get(LinkNotcia, headers=headers).text
                            Titulo = []
                            for Titu in confiTagPage["j"]["tituloNoticia"]:
                                try:
                                    Titulos = eval(Titu)
                                    if Titulos != []:
                                        Titulo.append(Titulos)
                                except Exception as e:
                                    # print(" Obtener Titulo", e,Titu)
                                    hoja.cell(row=row, column=17).value = str(e)
                            resultadoTitulo = contarElementosLista(Titulo)
                            if resultadoTitulo != {}:
                                maximoTitulo = max(resultadoTitulo, key=resultadoTitulo.get)
                                # hoja.cell(row=row, column=6).value = maximoTitulo
                                print("El valor mas repetido es el ", maximoTitulo, " con ",
                                      resultadoTitulo[maximoTitulo], " veces")

                            Descripcion = []
                            for Descr in confiTagPage["j"]["descripcionNoticia"]:
                                try:
                                    Descripciones = eval(Descr)
                                    if Descripciones != []:
                                        Descripcion.append(Descripciones)
                                except Exception as e:
                                    # print(" Obtener Titulo", e,Descr)
                                    hoja.cell(row=row, column=18).value = str(e)
                            resultadoDescripcion = contarElementosLista(Descripcion)
                            if resultadoDescripcion != {}:
                                maximoDescripcion = max(resultadoDescripcion, key=resultadoDescripcion.get)
                                # hoja.cell(row=row, column=7).value = maximoDescripcion
                                print("El valor mas repetido es el ", maximoDescripcion, " con ",
                                      resultadoDescripcion[maximoDescripcion], " veces")

                            link_web = url
                            FechaHoraScrapeo = str(datetime.datetime.now())
                            archivoCSV.append(link_web + ',\n' + ',\n' + LinkNotcia + ',\n' + FechaHoraScrapeo)

                            with open("out.csv", "a") as f:
                                wr = csv.writer(f, delimiter="\n")
                                for ele in items:
                                    wr.writerow([ele + ","])
                    except Exception as e:
                        print(" ERROR AL OBTENER TITULO O DESCRIPCION", e)

                    items.append(LinkNotcia)

                    ### store the
                    # texto2els(texto, FechaHoraScrapeo, LinkNotcia)
                    link = LinkNotcia

                    # resto 5 horas porque es la diferencia horario que tengo con Alemania
                    aive_dt = datetime.datetime.now() - - datetime.timedelta(hours=5)
                    fecha = aive_dt.strftime("%Y-%m-%dT%H:%M:%SZ")

                    titulo = maximoTitulo
                    # stract portal media from notice's link
                    medio = link2medio(link)
                    grupo = grupo_telegram_fijo

                    #store.store(titulo, fecha, texto, link, medio, grupo)

                    #########################################################################

            LINKS.close()
            PagNoticiaLink.save('./Excel/' + urlCortada + '-Noticias.xlsx')
            # hojaExcelDeErrores.save('./Excel/errores' + urlCortada + '-Errores.xlsx')
            return items
        except Exception as e:
            print(" 100 - Obtener links ", e)
            hoja.cell(row=row, column=10).value = str(e)
def limpiar(texto, mugre):
    for m in mugre:
        texto = texto.replace(m, "")
    return texto
def hashear(l):
    l = l.encode('utf-8')
    h = hashlib.new( "sha1",l)
    return h.hexdigest()
def link_enviado(l):

    print("Telegram analiza este link para enviar ",l)
    l = limpiar(l, mugre)
    l = hashear(l)

    if not l in j_link_enviado.keys():
        print(" Enviando a telegram :", l)
        j_link_enviado[l] = 1
        return False
    else:
        print(" !!!!!!!!!!!!!!! ENCONTRO DUPLICADO !!!!!!!!!!!!!!!!!! ")
        return True
def configuracion():
    f = open("configParametrosGenerico.json", "r")
    global j_config
    j_config = {}
    j_config = json.loads(f.read())
    global vtelegram
    global grupo_telegram_fijo
    grupo_telegram_fijo = j_config["grupo_telegram_fijo"]
    global url_api_sonda
    url_api_sonda = j_config["url_api_sonda"]
    try:
        vtelegram = j_config["telegram"]
    except:
        vtelegram = True
    global token
    token = j_config["token"]
    global idchat
    idchat = j_config["id_chat"]
    global Tema
    Tema = j_config["Tema"]
    global NombreGrupo
    NombreGrupo = j_config["NombreGrupo"]
    global directorio
    directorio = j_config["directorio"]
    try:
        global db_noticias2
        db_noticias2 = {}
        if os.path.isfile("./persist/db_noticias2"+id_grupo_telegram+".bin"):
            db_noticias2 = load_persist("./persist/db_noticias2"+id_grupo_telegram)
        else:
            db_noticias2 = {}
    except:
        db_noticias2 = {}
    try:
        global LinksDePaginasWeb
        LinksDePaginasWeb = {}
        if os.path.isfile(directorio + 'persist/LinksDePaginasWeb.bin'):
            LinksDePaginasWeb = load_persist("LinksDePaginasWeb")
        else:
            LinksDePaginasWeb = {}
    except:
        LinksDePaginasWeb = {}
def enviar_noticias(arr,id_chat,Nombre_Grupo,provincias,Temas):
    if not vtelegram:
        pass
        # return
    filtro_repetida

    try:
        url_api = token + "/sendMessage"

        # print( "- tema \n", Tema, " \n ",  NombreGrupo )
        men_t = "✔ Noticias referidas al tema %s, enviadas al grupo de télegram %s: " % (Temas, Nombre_Grupo) + "\n"
        ta = False
        # recorro el arreglo de links y lo imprimos
        men = []
        # print("arr \n",  arr)
        # print( "men_t \n", men_t )
        for a in arr:
            # print("3- ",a)
            # men += "- " + a + "\n\n"

            # armo la linea
            l = "- " + a + "\n\n"
            men.append(l)
            if a != "":
                ta = True
        if ta:
            # Si tiene información, mando el título.
            requests.post('https://api.telegram.org/' + url_api, data={'chat_id': id_chat, 'text': men_t})
            for m in men:

                # todo Eze, fijate que tuve que poner esta función par que controle no mandar repetidos.
                if not link_enviado(m):

                    requests.post('https://api.telegram.org/' + url_api,
                                  data={'chat_id': id_chat, 'text': '\n [' + Nombre_Grupo + ']\n' + m})
                    print(requests.status_codes)
    except Exception as e:
        print(" 279 - enviar ", e)

def procesar_id(k):
    # obtengo el valor de la sonda para el k
    valor_sonda = datos_sonda[k]
    id_telegram = k
    nombre_grupo = valor_sonda[0]
    global id_grupo_telegram
    id_grupo_telegram = k
    tema = valor_sonda[1]
    provincias = valor_sonda[2]
    json_donde_estan_los_portales = formar_json_portales(provincias)
    j = open("./configGenerico.json", "r")
    configuracion()
    confiTagPage = {}
    confiTagPage = json.loads(j.read())
    for url in json_donde_estan_los_portales:
        if json_donde_estan_los_portales != "":
            # print( "********************************" )
            print("TEMA:\n\n", tema)
            # print( "********************************" )
            print(" Procesando la url:  ", url)
            r = RSSParser().parse(confiTagPage, url, tema)
            PagNoticiaLink.save('./Excel/' + urlCortada + '-Noticias.xlsx')
            if r != []:
                enviar_noticias(r,id_telegram,nombre_grupo,provincias,tema)
    """
    acá va el programa tal cual está en el main original ahora
    
    """

### main ###
## supongamos que tenemos en memoria estos datos en la sonda
datos_sonda = {
    "-123456789": ("sb_019_Presidente", ["termino1,termino2"], ["BsAs", "Catamarca"]),
    "-123456781": ("sb_010_otro", ["termino1,termino2"], ["Neuquen", "Mendoza"]),
    "-123456782": ("sb_020_otro", ["termino1,termino2"], ["Cordoba", "Catamarca"]),
    "-123456783": ("sb_012_X", ["termino1,termino2"], ["Cordoba", "San_Luis"]),
    "-123456784": ("sb_014_Y", ["termino1,termino2"], ["Tierra_Del_Fuego", "La_Rioja"])
}
def verifico_cambios_en_la_sonda():
    # traigo los datos de las sonda

    url = "http://167.86.120.98:5003/sb_get_sonda2t/"
    url2 = "http://167.86.120.98:5003/sb_get_sonda2j/"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:55.0) Gecko/20100101 Firefox/55.0',
    }
    dict1 = {}
    # = requests.get(url, headers=headers).text
    datos_sonda_temp2 = requests.get(url2, headers=headers).json()
    datos_sonda_temp = requests.get(url, headers=headers)
    datos_sonda_temp = ast.literal_eval(datos_sonda_temp.content.decode('utf-8'))
    #datos_sonda_temp = {
     #   "-474068207": ("sb_023_Jujuy", ["jujuy,bolsa"], ["Jujuy", "Chaco"])
    #}
    global datos_sonda
    if datos_sonda != datos_sonda_temp:
        datos_sonda = datos_sonda_temp
if __name__ == '__main__':

    # tomo el primer parámetro que se le pasa al scrip cuando se ejecuta
    # en este caso, es el nro de terminación del id de telegram
    terminacion_id = sys.argv[1]

    # tiene que haber un bucle infinito
    while True:

        verifico_cambios_en_la_sonda()

        for k in datos_sonda:

            if corresponde_procesar_id(k, terminacion_id):
                procesar_id(k)
